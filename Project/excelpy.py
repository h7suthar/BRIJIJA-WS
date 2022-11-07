import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


# Defining and initialzing the data feild for the given dataset...
# Data is in the form of set of dictionary of various items...

data = {
    # Books...
    "Gita": {
        "Categaory": 'Books',
        "Seller": 'Valmiki',
        "Quantity": 10,
        "Price": 30.00,
        "Warranty": 'N/A',
        "Discount": 10,
    },
    "Ramayan": {
        "Categaory": 'Books',
        "Seller": 'Valmiki',
        "Quantity": 10,
        "Price": 25.00,
        "Warranty": 'N/A',
        "Discount": 15
    },
    "Ikigai": {
        "Categaory": 'Books',
        "Seller": 'Mitsuhshi',
        "Quantity": 10,
        "Price": 35.00,
        "Warranty": 'N/A',
        "Discount": 5
    },
    "Atomic-Habits": {
        "Categaory": 'Books',
        "Seller": 'Pearson',
        "Quantity": 10,
        "Price": 40.00,
        "Warranty": 'N/A',
        "Discount": 0
    },
    "RDad-PDad": {
        "Categaory": 'Books',
        "Seller": 'Gallery',
        "Quantity": 10,
        "Price": 25.00,
        "Warranty": 'N/A',
        "Discount": 2
    },
    # Electronic...
    "Television": {
        "Categaory": 'Electronics',
        "Seller": 'Sony',
        "Quantity": 5,
        "Price": 599.00,
        "Warranty": '1 Year',
        "Discount": 10
    },
    "Laptop": {
        "Categaory": 'Electronics',
        "Seller": 'HP',
        "Quantity": 5,
        "Price": 999.00,
        "Warranty": '15 Months',
        "Discount": 20
    },
    "Speakers": {
        "Categaory": 'Electronics',
        "Seller": 'Bose',
        "Quantity": 5,
        "Price": 299.00,
        "Warranty": '1 Year',
        "Discount": 0
    },
    "Camera": {
        "Categaory": 'Electronics',
        "Seller": 'Canon',
        "Quantity": 5,
        "Price": 599.00,
        "Warranty": '1 Year',
        "Discount": 7.5
    },
    "Printer": {
        "Categaory": 'Electronics',
        "Seller": 'Epson',
        "Quantity": 5,
        "Price": 349.00,
        "Warranty": '1 Year',
        "Discount": 0
    },
    # Grocery...
    "Wheat-FLour": {
        "Categaory": 'Grocery',
        "Seller": 'Sujata',
        "Quantity": 24,
        "Price": 9.99,
        "Warranty": 'N/A',
        "Discount": 0
    },
    "Rice": {
        "Categaory": 'Grocery',
        "Seller": 'Royal',
        "Quantity": 24,
        "Price": 14.99,
        "Warranty": 'N/A',
        "Discount": 0
    },
    "Black-Beans": {
        "Categaory": 'Grocery',
        "Seller": 'Goya',
        "Quantity": 24,
        "Price": 3.99,
        "Warranty": 'N/A',
        "Discount": 0
    },
    "Musturd-Seeds": {
        "Categaory": 'Grocery',
        "Seller": 'Deep',
        "Quantity": 24,
        "Price": 4.99,
        "Warranty": 'N/A',
        "Discount": 0
    },
    "Almonds": {
        "Categaory": 'Grocery',
        "Seller": 'Cali-Almd',
        "Quantity": 24,
        "Price": 12.99,
        "Warranty": 'N/A',
        "Discount": 0
    },
    # Sports...
    "Bat": {
        "Categaory": 'Sports',
        "Seller": 'Kukaburra',
        "Quantity": 12,
        "Price": 249.00,
        "Warranty": '2 Months',
        "Discount": 0
    },
    "Football": {
        "Categaory": 'Sports',
        "Seller": 'Puma',
        "Quantity": 12,
        "Price": 99.00,
        "Warranty": '3 Months',
        "Discount": 0
    },
    "Basketball": {
        "Categaory": 'Sports',
        "Seller": 'Splading',
        "Quantity": 12,
        "Price": 129.00,
        "Warranty": '3 Months',
        "Discount": 0
    },
    "Tennis-Racket": {
        "Categaory": 'Sports',
        "Seller": 'Wilson',
        "Quantity": 12,
        "Price": 49.00,
        "Warranty": '30 Days',
        "Discount": 0
    },
    "Golf-Stick": {
        "Categaory": 'Sports',
        "Seller": 'Under-Armour',
        "Quantity": 12,
        "Price": 199.00,
        "Warranty": '6 Months',
        "Discount": 0
    },
    # Wardrobe...
    "Suit": {
        "Categaory": 'Wardrobe',
        "Seller": 'Hagger',
        "Quantity": 6,
        "Price": 149.00,
        "Warranty": '6 Months',
        "Discount": 20
    },
    "Shirt": {
        "Categaory": 'Wardrobe',
        "Seller": 'Van-Hausen',
        "Quantity": 6,
        "Price": 49.00,
        "Warranty": '3 Months',
        "Discount": 10
    },
    "Pant": {
        "Categaory": 'Wardrobe',
        "Seller": 'Dockers',
        "Quantity": 6,
        "Price": 69.00,
        "Warranty": '1 Months',
        "Discount": 15
    },
    "Polo": {
        "Categaory": 'Wardrobe',
        "Seller": 'Nautica',
        "Quantity": 6,
        "Price": 59.00,
        "Warranty": 'None',
        "Discount": 0
    },
    "Gawn": {
        "Categaory": 'Wardrobe',
        "Seller": 'Flora',
        "Quantity": 6,
        "Price": 99.00,
        "Warranty": 'None',
        "Discount": 25
    }
}

# Methods for Initizing the dataset...
def get_finalPrice(discount,orig_Price):
    disAmount = ((orig_Price*discount) / 100)
    final_Price = (orig_Price - disAmount)
    return final_Price

def check_file(filename):
    if os.path.exists(filename):
        return True
    else:
        return False

def getHeader_Index(worksheet,header):
    for cell in worksheet['1']:
        if cell.value == header:
            column = get_column_letter(cell.column)
    return column

def initialize_workbook(data_Entry):

    # Checking if workbook is already ceated...
    file = check_file('NewInventry.xlsx')
    if not file:

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"

        item_Details = ['Item','Categaory','Seller','Quantity','Original Price','Warranty','Discount','Final Price']
        ws.append(item_Details)

        # Inserting the data into the Excel sheet...
        for item in data_Entry:
            details = list(data_Entry[item].values())
            ws.append([item] + details)

        # Heading for item details...
        for column in range(1,len(item_Details)+2):
            ws[get_column_letter(column) + '1'].font = Font(bold=True)

        # Finding the cell column index for original price, discount and final price...
        for cell in ws['1']:
            if cell.value == 'Original Price':
                column = get_column_letter(cell.column)
                OP = column
            if cell.value == 'Discount':
                column = get_column_letter(cell.column)
                DIC = column 
            if cell.value == 'Final Price':
                column = get_column_letter(cell.column)
                FP = column 

        # Calculating the final price after the discount and loading it in to excel column...
        for row in range(2,(ws.max_row + 1)):
            origPrice = ws[OP + str(row)].value
            discPercent = ws[DIC + str(row)].value
            ws[FP + str(row)].value = get_finalPrice(discPercent,origPrice)

        # Saving the Initialized workbook...
        wb.save("NewInventry.xlsx")
        return ws,wb

    # If created just load it...
    else:
        wb = load_workbook('NewInventry.xlsx')
        ws = wb.active
        return ws,wb


# Initializing the Dataset...
database,dataBook = initialize_workbook(data)